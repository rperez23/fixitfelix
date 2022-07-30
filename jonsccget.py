#! /usr/bin/python3

#we will have already downloaded the scc files BUZ_######.scc
#this script will rename to show_s#_e#_fc.scc
#currently 1,174 scc files in s3://s3-fremantle-ccfiles-or-1/Caption Uploads/FAST CHANNELS/

import os
import re
import sys
import time
import argparse
import openpyxl
import warnings
import subprocess
import shutil
from openpyxl.utils.cell import get_column_letter

tab = "1. Master Metadata"

warnings.simplefilter(action='ignore', category=UserWarning)

#done
def getXLF():

    print("")
    while True:
        s3path = input("  Give me your AWS S3 Path: ")

        if s3path == "":
            sys.exit(0)

        lscmd  = "aws s3 ls \"" + s3path + "\" | grep xlsx | grep -v \"(\"| grep -v \"S10-DH\" | grep -v _km | grep -v TEST | awk '{print $4}'"
        #lscmd  = "aws s3 ls \"" + s3path + "\" | grep xlsx | awk '{print $4}'"
        xlf = os.popen(lscmd).read()
        time.sleep(5)
        xlf = xlf.strip()

        s3m = re.search("FAST_CHANNEL/$",s3path)
        #s3m = re.search("_EditedMaster",s3path)
        xlfm = re.search('\.xlsx$',xlf)

        if s3m and xlfm:
            break

    cpcmd = "aws s3 cp \"" + s3path + xlf + "\" ."
    print(" ",cpcmd)
    cpoutput = os.popen(cpcmd)
    time.sleep(5)

    if not os.path.exists(xlf):
        print("  ~~Could not copy xlf file")
        sys.exit(1)



    return xlf, s3path

#done
def openXLF(xlf):

    try:
        workbook = openpyxl.load_workbook(filename=xlf,read_only=True)
    except:
        print("  Cannot open",xlf)
        sys.exit(0)

    #validate that sheet exists
    if not (tab in workbook.sheetnames):
        print(" ",tab,"not in",xlf)
        sys.exit(0)

    #assign the work sheet object
    ws = workbook[tab]

    return workbook, ws

#done
def archiveCaption(sccname,housenum):

    subdirList = ['Episode_Number', 'House_Number']

    capmatch = re.search("^([a-zA-Z]+)_.+_([sS]\d+)_",sccname)

    if capmatch:
        show   = capmatch.group(1)
        season = capmatch.group(2)

        season   = season.upper()
        dirname  = show + "_" + season

        for subdir in subdirList:
            fullpath = "archive/" + dirname + "/" + subdir

            if not os.path.exists(fullpath):
                try:
                    os.makedirs(fullpath)
                except:
                    print("  ~~Could not create",fullpath)
                    return 1


        try:
            os.link(housenum,"archive/" + dirname + "/" + "House_Number/" + housenum)
        except:
            print("  ~~Could not link to House_Number:", housenum)
            sys.exit(1)

        try:
            os.link(housenum,"archive/" + dirname + "/" + "Episode_Number/" + sccname)
        except:
            print("  ~~Could not link to Episode_Number:", housenum)
            sys.exit(1)

    return 0


#done
def archiveSubs(ws,subcol,hncol):

    r = 6
    while True:

        c  = subcol
        sccname  = str(ws.cell(r,c).value)

        c = hncol
        housenum = str(ws.cell(r,c).value)

        if (sccname == "None") or (housenum == "None"):
            return
        else:
            sccname  += ".scc"
            housenum += ".scc"
            #print(sccname, ":", housenum)
            archiveCaption(sccname,housenum)
            shutil.move(housenum,"zDone/captions/" + housenum)

        r += 1


#done: will report columns of scc & house number
def getColumns(ws):

    row      = 5
    col      = 2
    endcol   = 34
    hncol    = 0
    subcol   = 0

    while (subcol == 0) or (hncol == 0):

        txt = str(ws.cell(row,col).value)
        txt = txt.lower()

        if txt == "scc filename \n(no extension)":
            subcol = col
        elif txt == "house number":
            hncol = col
        elif txt == "none":
            print("  Could not find columns")
            sys.exit(1)

        col += 1

    #print("SCC :",subcol," HN:",hncol)
    return subcol,hncol



#Main Program
while True:
    #1) Get the XLF File
    xlf,s3path = getXLF()

    #2) open the XLF File: get workbook and worksheet
    workbook,ws = openXLF(xlf)

    #3) get the columns of the scc files, and HouseNumbers:
    subcol, hncol = getColumns(ws)

    #4) read the xl file and archive the data
    archiveSubs(ws,subcol,hncol)

    workbook.close()
    shutil.move(xlf,"zDone/" + xlf)
    print("=======================")
