#! /usr/bin/python3

import os
import re
import sys
import time
import argparse
import openpyxl
import warnings
import subprocess
import webbrowser
from openpyxl.utils.cell import get_column_letter

tab = "1. Master Metadata"

warnings.simplefilter(action='ignore', category=UserWarning)

def getXLF():

    print("")
    while True:
        s3path = input("  Give me your AWS S3 Path: ")
        if s3path == "":
            sys.exit(0)
        lscmd  = "aws s3 ls \"" + s3path + "\" | grep xlsx | awk '{print $4}'"
        xlf = os.popen(lscmd).read()
        time.sleep(5)
        xlf = xlf.strip()

        s3m = re.search("FAST_CHANNEL/$",s3path)
        #s3m = re.search("- Wurl/$",s3path)
        xlfm = re.search('\.xlsx$',xlf)

        if s3m and xlfm:
            break

    cpcmd = "aws s3 cp \"" + s3path + xlf + "\" ."
    print(" ",cpcmd)
    cpoutput = os.popen(cpcmd)
    time.sleep(5)

    if not os.path.exists(xlf):
        print("  ~~Could not copy xlf file")
        sys.exit(1)



    return xlf, s3path

def openXLF(xlf):

    try:
        workbook = openpyxl.load_workbook(filename=xlf)
    except:
        print("  Cannot open",xlf)
        sys.exit(0)

    #validate that sheet exists
    if not (tab in workbook.sheetnames):
        print(" ",tab,"not in",xlf)
        sys.exit(0)

    #assign the work sheet object
    ws = workbook[tab]

    return workbook, ws

def makeBackup(xlf):
    cpcmd = "cp " + xlf + " 0rig"
    #print("    making backup file")
    os.popen(cpcmd)
    time.sleep(5)
    backupf = "0rig/" + xlf
    if not os.path.exists(backupf):
        print("  ~~Could not make backup file~~")
        sys.exit(1)

def getFormatVal():
    formatval = input("  What is the Format for this Show : ")
    return formatval

def editXlf(ws,formatval):

    startrow = 5
    r        = startrow + 1
    startcol = 2
    endcol   = 34

    #formatval = input("  What is the Format for this Show : ")

    for c in range(startcol,endcol + 1):

        txt = str(ws.cell(startrow,c).value)
        txt = txt.lower()
        #print(txt)
        if txt == "format":
            break
        elif txt == "none":
            sys.exit(1)

    while True:

        txt = str(ws.cell(r,c).value)
        txt = txt.lower()
        if txt == "none":
            break
        else:
            ws.cell(row=r, column=c).value = formatval
        r += 1

def sendItBack(xlf,s3path):

    s3 = "\"" + s3path + "\""
    cpcmd = "aws s3 cp " + xlf + " " + s3
    print(" ",cpcmd)
    cpoutput = os.popen(cpcmd)
    time.sleep(5)
    print(cpoutput)

def getIdLink(ws):

    base = "https://core.wazeedigital.com/video/searchResults.do?search.type=intermediate&search.withinKeywords=&search.withinResults=&filter=v1:"
    hns  = ""

    startrow = 5
    r        = startrow + 1
    startcol = 2
    endcol   = 34


    for c in range(startcol,endcol + 1):
        txt = str(ws.cell(startrow,c).value)
        txt = txt.lower()

        if txt == "house number":
            print("  House Number")
            print("  ============")
            break
        elif txt == "none":
            sys.exit(1)


    while True:

        txt = str(ws.cell(r,c).value)

        if txt == "None":
            m = re.search("(.+)%20OR%20$",hns)
            if m:
                tot = (r - 1) - (startrow + 1) + 1
                print("\n ",tot,"House Numbers Found\n")
                zelda = base + m.group(1)
                webbrowser.open(zelda)
            break
        else:
            print(" ",txt)
            hns += "Fremantle.HouseNumber:" + txt + "%20OR%20"

        r += 1


formatval = getFormatVal()

while True:

    #1) Get the XLF File
    xlf,s3path = getXLF()

    #2) copy the XLF to the 0rig directory
    makeBackup(xlf)

    #3) open the XLF File: get workbook and worksheet
    workbook,ws = openXLF(xlf)

    #4) get Buzzer Numbers (maybe print link)
    getIdLink(ws)

    #5) edit the xl file
    editXlf(ws,formatval)

    #6) close the workbook
    workbook.save(xlf)
    workbook.close()

    # move it to done folder
    try:
        os.rename(xlf,"zDone/" + xlf)
    except:
        print("  ~~Cannot move xlf to zDone:",xlf)
        sys.exit(0)

    #7) Copy it back to aws
    sendItBack("zDone/" + xlf,s3path)
    print("\n=======================\n")
